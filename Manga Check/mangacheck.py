#THIS PROGRAM WAS CREATED BY NOVER SAFERT

import logging
from openpyxl import Workbook 
from openpyxl import load_workbook
import webbrowser
wb = Workbook()

def xlsx_loader():
    wb1 = load_workbook('Mangacheck.xlsx')
    logging.debug('xlsx loaded successfully')
    logging.debug( wb1.sheetnames)
    return (wb1)

def checkall(ws,size):
    size = size + 2
    for num in range(2, size):
        cell = "A" + str(num)
        print(ws[cell].value)

def main():
    logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
    logging.debug('Start of program')
    wb1 = xlsx_loader()
    ws = wb1['Mangas']
    option = 0
    size = 12
    while (option !=  -1):
        print(" [1] Check all manga\n [2] Add a manga")
        print(" [3] Edit manga info \n [-1] Exit\n")
        option = int(input("Select the number of the option you want\n"))
        if (option == 1):
            checkall(ws, size)
            print("Do you want to any thing else?")
            op2 = str(input("[Y] [N]\n"))
            if (op2 == "N" or "n"):
                break

main()